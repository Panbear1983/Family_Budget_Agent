#!/usr/bin/env python3
"""
Cell-by-Cell Budget Editor
Allows manual editing of individual cells with preview and validation
"""

import sys
import os
import re
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
from rich.console import Console
from rich.table import Table
from rich import box as rich_box
from datetime import datetime, timedelta
import calendar
import config
import openpyxl
from utils.excel_totals import save_workbook_with_totals

EXCEL_FILE_PATH = config.BUDGET_PATH

def get_active_year() -> int:
    """
    Infer the year from the active workbook filename.
    Expected: "{year}年開銷表（NT）.xlsx"
    """
    basename = os.path.basename(EXCEL_FILE_PATH)
    m = re.search(r'(\d{4})年開銷表', basename)
    if m:
        return int(m.group(1))
    return getattr(config, "CURRENT_YEAR", datetime.now().year)

def get_days_in_month(month_name):
    """Get number of days in a month"""
    month_map = {
        '一月': 1, '二月': 2, '三月': 3, '四月': 4, '五月': 5, '六月': 6,
        '七月': 7, '八月': 8, '九月': 9, '十月': 10, '十一月': 11, '十二月': 12
    }
    month_num = month_map.get(month_name)
    if month_num:
        year = get_active_year()
        return calendar.monthrange(year, month_num)[1]
    return 31  # Default

def get_month_number(month_name):
    """Convert month name to number"""
    month_map = {
        '一月': 1, '二月': 2, '三月': 3, '四月': 4, '五月': 5, '六月': 6,
        '七月': 7, '八月': 8, '九月': 9, '十月': 10, '十一月': 11, '十二月': 12
    }
    return month_map.get(month_name, 1)

def select_month():
    """Let user select a month to edit - shows preview before confirming"""
    console = Console()
    months = ['一月', '二月', '三月', '四月', '五月', '六月',
              '七月', '八月', '九月', '十月', '十一月', '十二月']
    
    console.print("\n[bold cyan]選擇月份 (Select Month):[/bold cyan]\n")
    for i, month in enumerate(months, 1):
        days = get_days_in_month(month)
        console.print(f"   [[green]{i:2d}[/green]] {month} ({days}天)")
    
    console.print(f"\n   [[green] x[/green]] 返回 (Back)")
    
    choice = input("\n選擇 (Choose): ").strip()
    
    if choice == 'x':
        return None
    
    try:
        month_num = int(choice)
        if 1 <= month_num <= 12:
            selected_month = months[month_num - 1]
            
            # Import and show current state of the month
            from utils.view_sheets import display_monthly_sheet
            
            console.print("\n" + "="*100)
            console.print(f"  📄 {selected_month} - Current State".center(100))
            console.print("="*100 + "\n")
            
            display_monthly_sheet(selected_month)
            
            console.print("\n" + "="*100 + "\n")
            
            # Ask for confirmation
            confirm = input("Do you want to edit this month? [y/N]: ").strip().lower()
            
            if confirm == 'y':
                return selected_month
            else:
                console.print("[yellow]Cancelled - returning to month selection[/yellow]")
                return None
                
    except ValueError:
        pass
    
    console.print("[red]Invalid choice[/red]")
    return None

def autofill_dates_workflow(sheet_name):
    """Workflow for auto-filling dates in a month"""
    console = Console()
    
    month_num = get_month_number(sheet_name)
    days_in_month = get_days_in_month(sheet_name)
    
    console.print(f"\n📅 [bold green]自動填充 {sheet_name} 日期 (Auto-fill {sheet_name} Dates)[/bold green]")
    console.print("="*100 + "\n")
    console.print(f"本月天數：{days_in_month} 天 (This month has {days_in_month} days)\n")
    
    console.print("[cyan]禮拜幾是本月第一天？(Which day of the week is the 1st?)[/cyan]\n")
    weekdays = [
        "禮拜一 (Monday)",
        "禮拜二 (Tuesday)", 
        "禮拜三 (Wednesday)",
        "禮拜四 (Thursday)",
        "禮拜五 (Friday)",
        "禮拜六 (Saturday)",
        "禮拜日 (Sunday)"
    ]
    
    for i, day in enumerate(weekdays, 1):
        console.print(f"   [[green]{i}[/green]] {day}")
    
    console.print("\n   [[green]x[/green]] 取消 (Cancel)")
    
    choice = input("\n選擇 (Choose): ").strip()
    
    if choice == 'x':
        return
    
    try:
        start_weekday = int(choice)
        if not (1 <= start_weekday <= 7):
            console.print("[red]無效選擇 (Invalid choice)[/red]")
            return
    except ValueError:
        console.print("[red]無效選擇 (Invalid choice)[/red]")
        return
    
    # Generate date preview
    year = get_active_year()
    console.print(f"\n✓ 已選擇: {weekdays[start_weekday - 1]}")
    console.print(f"📆 月份: {sheet_name} ({year}年)")
    console.print(f"📊 本月天數: {days_in_month} 天\n")
    
    # Calculate dates
    first_date = datetime(year, month_num, 1)
    
    # Define the weekly blocks (row indices)
    week_blocks = [
        (3, 9),    # Week 1: rows 3-9 (0-indexed: 2-8)
        (11, 17),  # Week 2: rows 11-17 (0-indexed: 10-16)
        (19, 25),  # Week 3
        (27, 33),  # Week 4
        (35, 41),  # Week 5
        (43, 49),  # Week 6
    ]
    
    # Generate all dates for the month
    dates_to_fill = []
    current_date = first_date
    
    # Monday = 1, Tuesday = 2, ..., Sunday = 7 (our input)
    # But we need to map to week blocks where each block is Mon-Sun
    # Row indices within each block: 0=Mon, 1=Tue, 2=Wed, 3=Thu, 4=Fri, 5=Sat, 6=Sun
    
    # Start from the first date - figure out which week and which day
    current_weekday = start_weekday  # 1-7 (Mon-Sun)
    
    week_idx = 0
    day_idx = current_weekday - 1  # 0-6 (Mon=0, Sun=6)
    
    for day_num in range(1, days_in_month + 1):
        # FIX: Use datetime object instead of string for better data typing in Excel
        current_date = datetime(year, month_num, day_num)
        
        if week_idx < len(week_blocks):
            start_row, end_row = week_blocks[week_idx]
            # Convert to 0-indexed
            row_idx = start_row - 1 + day_idx
            
            dates_to_fill.append({
                'row': row_idx,
                'col': 0,  # Column A
                'value': current_date,
                'display': f"Row {row_idx + 1} (Week {week_idx + 1}, Day {day_idx + 1})"
            })
        
        # Move to next day
        day_idx += 1
        if day_idx >= 7:  # New week
            day_idx = 0
            week_idx += 1
    
    # Show detailed preview with yellow text
    console.print("\n[bold yellow]📋 預覽將要填充的日期 (Preview - All dates to be filled):[/bold yellow]\n")
    
    weekday_names = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    weekday_cn = ["一", "二", "三", "四", "五", "六", "日"]
    
    # Group by week for better readability
    current_week = 0
    week_dates = []
    
    for date_info in dates_to_fill:
        week_num = (date_info['row'] - 2) // 8  # Calculate week number
        
        if week_num != current_week:
            # Print previous week if exists
            if week_dates:
                console.print(f"\n  [cyan]第{current_week + 1}週:[/cyan]")
                for d in week_dates:
                    day_idx = (d['row'] - 2) % 8
                    if 0 <= day_idx < 7:
                        weekday_label = f"{weekday_cn[day_idx]}"
                        console.print(f"    [yellow]{d['value']}[/yellow] (禮拜{weekday_label}) → Row {d['row'] + 1}")
                week_dates = []
            current_week = week_num
        
        week_dates.append(date_info)
    
    # Print last week
    if week_dates:
        console.print(f"\n  [cyan]第{current_week + 1}週:[/cyan]")
        for d in week_dates:
            day_idx = (d['row'] - 2) % 8
            if 0 <= day_idx < 7:
                weekday_label = f"{weekday_cn[day_idx]}"
                console.print(f"    [yellow]{d['value']}[/yellow] (禮拜{weekday_label}) → Row {d['row'] + 1}")
    
    console.print(f"\n[bold]總共將填充 {len(dates_to_fill)} 個日期[/bold]")
    console.print("[yellow]所有新添加的日期將用黃色標記在Excel中[/yellow]\n")
    
    # Confirm
    confirm = input("\n確認填充？(Confirm auto-fill?) [y/N]: ").strip().lower()
    
    if confirm != 'y':
        console.print("[yellow]已取消 (Cancelled)[/yellow]")
        return
    
    # Apply dates
    apply_dates(sheet_name, dates_to_fill)

def apply_dates(sheet_name, dates_to_fill):
    """Apply date fills to Excel file"""
    console = Console()
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb[sheet_name]
        
        # Apply dates
        for date_info in dates_to_fill:
            excel_row = date_info['row'] + 1  # Convert to 1-based
            excel_col = date_info['col'] + 1
            
            cell = ws.cell(row=excel_row, column=excel_col)
            cell.value = date_info['value']
        
        # Save
        wb.save(EXCEL_FILE_PATH)
        wb.close()
        
        console.print(f"\n[bold green]✅ 成功填充 {len(dates_to_fill)} 個日期！(Successfully filled {len(dates_to_fill)} dates!)[/bold green]\n")
        
    except Exception as e:
        console.print(f"\n[bold red]保存失敗 (Save failed): {e}[/bold red]\n")

# REMOVED: edit_category_labels_workflow() and apply_category_label()
# Reason: Prevents conflicts with hardcoded category mappings in categorizer and merger
# Category labels should be set in the template file, not changed mid-year

def edit_expenses_workflow(sheet_name):
    """Workflow for editing expenses with date + category selection"""
    console = Console()
    
    console.print(f"\n💰 [bold green]編輯支出 (Edit Expenses)[/bold green]")
    console.print("="*100 + "\n")
    
    try:
        # Read the sheet
        df = pd.read_excel(EXCEL_FILE_PATH, sheet_name=sheet_name, header=None)
        active_year = get_active_year()
        
        # Get category labels from Row 2, Columns C-I
        row_2 = df.iloc[1]
        categories = []
        for col_idx in range(3, 9):  # Columns D-I (indices 3-8)
            if col_idx < len(row_2):
                label = str(row_2.iloc[col_idx]) if pd.notna(row_2.iloc[col_idx]) else f'Category {col_idx - 1}'
                categories.append(label)
            else:
                categories.append(f'Category {col_idx - 1}')
        
        # Collect multiple expense edits
        edits = []
        
        while True:
            console.print("\n[cyan]輸入支出資料 (Enter expense data)[/cyan]")
            console.print("輸入 'done' 完成編輯 / Enter 'done' to finish\n")
            
            # Prompt for date
            date_input = input(f"輸入日期 (Date, e.g., {active_year}-10-15): ").strip()
            
            if date_input.lower() == 'done':
                break
            
            if date_input.lower() == 'cancel':
                return
            
            # Validate date format
            try:
                parsed_dt = datetime.strptime(date_input, "%Y-%m-%d")
            except:
                console.print("[red]日期格式錯誤，請使用 YYYY-MM-DD 格式 (Invalid date format)[/red]")
                continue

            if parsed_dt.year != active_year:
                console.print(f"[red]日期年份需為 {active_year}（目前編輯的是該年份的表）[/red]")
                continue
            
            # Find the row with this date in column A
            date_row = None
            for idx, row in df.iterrows():
                cell_val = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ''
                if date_input in cell_val:
                    date_row = idx
                    break
            
            if date_row is None:
                console.print(f"[red]找不到日期 {date_input} (Date not found)[/red]")
                continue
            
            # Display category menu
            console.print(f"\n✓ 找到日期: {date_input} (Row {date_row + 1})")
            console.print("\n選擇類別 (Select Category):\n")
            for i, cat in enumerate(categories, 1):
                console.print(f"   {i}. {cat}")
            
            cat_choice = input("\n選擇 (Choose): ").strip()
            
            try:
                cat_num = int(cat_choice)
                if not (1 <= cat_num <= 6):
                    console.print("[red]無效選擇 (Invalid choice)[/red]")
                    continue
            except ValueError:
                console.print("[red]無效選擇 (Invalid choice)[/red]")
                continue
            
            # Prompt for amount
            amount_input = input("\n輸入金額 (Amount): ").strip()
            
            try:
                amount = float(amount_input)
            except ValueError:
                console.print("[red]金額格式錯誤 (Invalid amount)[/red]")
                continue
            
            # Calculate column index (C=2, D=3, E=4, F=5, G=6, H=7, I=8)
            col_idx = 3 + (cat_num - 1)
            
            # Get old value
            old_val = df.iloc[date_row, col_idx] if pd.notna(df.iloc[date_row, col_idx]) else 0
            
            # Store edit
            edits.append({
                'row': date_row,
                'col': col_idx,
                'value': amount,
                'date': date_input,
                'category': categories[cat_num - 1],
                'old_value': old_val
            })
            
            console.print(f"[green]✓ 已添加: {date_input} - {categories[cat_num - 1]} = {amount}[/green]")
            console.print("[dim]─── 繼續輸入下一筆，或輸入 'done' 完成 (Next entry or 'done' to finish) ───[/dim]")
        
        if not edits:
            console.print("[yellow]未進行任何編輯 (No edits made)[/yellow]")
            return
        
        # Show summary of edits
        console.print(f"\n[bold]更改摘要 (Changes Summary):[/bold]")
        for edit in edits:
            row_display = edit['row'] + 1
            console.print(f"  [yellow]●[/yellow] Row {row_display} ({edit['date']}), {edit['category']}: [yellow]{edit['value']}[/yellow]")
        
        console.print(f"\n總共 [bold yellow]{len(edits)}[/bold yellow] 個單元格將被修改")
        
        # Confirm
        confirm = input("\n確認保存更改？(Confirm?) [y/N]: ").strip().lower()
        
        if confirm != 'y':
            console.print("[yellow]已取消 (Cancelled)[/yellow]")
            return
        
        # Apply edits
        apply_expense_edits(sheet_name, edits)
        
    except Exception as e:
        console.print(f"[red]錯誤 (Error): {e}[/red]")

def apply_expense_edits(sheet_name, edits):
    """Apply expense edits to Excel file"""
    console = Console()
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
        ws = wb[sheet_name]
        
        # Apply edits
        for edit in edits:
            excel_row = edit['row'] + 1  # Convert to 1-based
            excel_col = edit['col'] + 1
            
            cell = ws.cell(row=excel_row, column=excel_col)
            cell.value = edit['value']
        
        save_workbook_with_totals(wb, EXCEL_FILE_PATH, sheet_names=[sheet_name])
        wb.close()
        
        console.print(f"\n[bold green]✅ 支出已保存！已編輯 {len(edits)} 筆資料 (Expenses saved! {len(edits)} entries edited)[/bold green]\n")
        
    except Exception as e:
        console.print(f"\n[bold red]保存失敗 (Save failed): {e}[/bold red]\n")

def main():
    """Main entry point for cell-by-cell editor
    
    Note: Date auto-fill is now handled automatically when creating new year files.
    The autofill_dates_workflow() function is kept for manual re-filling if needed,
    but is not exposed in the normal menu flow.
    """
    console = Console()
    
    # Select month
    month = select_month()
    if not month:
        return
    
    # Go directly to expense editing (dates are auto-filled on year creation)
    edit_expenses_workflow(month)

if __name__ == "__main__":
    main()

