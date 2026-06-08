"""Recalculate daily, weekly, and monthly totals in budget Excel sheets.

Strategy:
- Write computed plain numbers to total cells so Python/pandas can always read them.
- App views also sum categories D-I directly (not dependent on column K).
- Call recalculate_month_totals() after any change to expense cells D-I.
"""

import config
import openpyxl

WEEKLY_BLOCKS = [
    (3, 9), (11, 17), (19, 25), (27, 33), (35, 41), (43, 49),
]
AMOUNT_COLS = [4, 5, 6, 7, 8, 9]  # D-I
TOTAL_COL = config.EXCEL_STRUCTURE["columns"]["daily_total"]  # Column K


def _amount(value):
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0


def recalculate_month_totals(ws):
    """Write plain-number totals for daily (col K), weekly, and monthly rows."""
    for start_row, end_row in WEEKLY_BLOCKS:
        for row_num in range(start_row, end_row + 1):
            row_sum = sum(_amount(ws.cell(row_num, col).value) for col in AMOUNT_COLS)
            ws.cell(row_num, TOTAL_COL, row_sum if row_sum > 0 else None)

    for start_row, end_row in WEEKLY_BLOCKS:
        total_row = end_row + 1
        week_total = 0
        for col in AMOUNT_COLS:
            week_sum = sum(_amount(ws.cell(r, col).value) for r in range(start_row, end_row + 1))
            ws.cell(total_row, col, week_sum if week_sum > 0 else None)
            week_total += week_sum
        ws.cell(total_row, TOTAL_COL, week_total if week_total > 0 else None)

    for row_num in range(50, 62):
        cell_val = str(ws.cell(row_num, 1).value or '')
        if '單項總額' in cell_val or '单项总额' in cell_val:
            month_total = 0
            for col in AMOUNT_COLS:
                month_sum = sum(
                    _amount(ws.cell(r, col).value)
                    for s, e in WEEKLY_BLOCKS
                    for r in range(s, e + 1)
                )
                ws.cell(row_num, col, month_sum if month_sum > 0 else None)
                month_total += month_sum
            ws.cell(row_num, TOTAL_COL, month_total if month_total > 0 else None)
            break


def repair_all_month_totals(file_path):
    """Recalculate totals for every month sheet in the workbook."""
    wb = openpyxl.load_workbook(file_path)
    months = config.EXCEL_STRUCTURE["month_sheets"]
    repaired = []
    for month in months:
        if month not in wb.sheetnames:
            continue
        recalculate_month_totals(wb[month])
        repaired.append(month)
    wb.calculation.calcMode = 'auto'
    wb.save(file_path)
    wb.close()
    return repaired


def save_workbook_with_totals(wb, file_path, sheet_names=None):
    """Save workbook after refreshing totals on the touched month sheet(s)."""
    months = sheet_names or config.EXCEL_STRUCTURE["month_sheets"]
    for month in months:
        if month in wb.sheetnames:
            recalculate_month_totals(wb[month])
    wb.calculation.calcMode = 'auto'
    wb.save(file_path)
