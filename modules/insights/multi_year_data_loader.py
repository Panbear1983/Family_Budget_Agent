"""
Multi-Year Data Loader - Load and merge data from multiple years
Extends DataLoader to support continuous timeline analysis across years
"""

import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple
from datetime import datetime, timedelta
from .data_loader import DataLoader


class MultiYearDataLoader(DataLoader):
    """Loads and merges data from multiple budget files"""
    
    def __init__(self, budget_files: List[str]):
        """
        Initialize multi-year data loader
        
        Args:
            budget_files: List of file paths [2025年開銷表.xlsx, 2026年開銷表.xlsx]
        """
        self.budget_files = budget_files
        self.cache = {}
        self.last_loaded = None
        self.ttl = 1800  # Cache for 30 minutes
        self.use_rolling_window = True  # Enable rolling 12-month window by default

        # Month normalization helpers (Chinese, English, numeric)
        self._month_names = ['一月', '二月', '三月', '四月', '五月', '六月',
                             '七月', '八月', '九月', '十月', '十一月', '十二月']
        self._english_month_map = {
            'january': '一月', 'february': '二月', 'march': '三月', 'april': '四月',
            'may': '五月', 'june': '六月', 'july': '七月', 'august': '八月',
            'september': '九月', 'october': '十月', 'november': '十一月', 'december': '十二月'
        }
        
        # Extract years from filenames
        self.years = []
        for file in budget_files:
            # Extract year from filename like "2025年開銷表（NT）.xlsx"
            import os
            filename = os.path.basename(file)
            year = filename[:4]
            if year.isdigit():
                self.years.append(int(year))
        
        self.years.sort()
        print(f"  📅 Multi-Year Loader: {len(self.years)} years ({', '.join(map(str, self.years))})")
        if self.use_rolling_window:
            print(f"  📊 Using rolling 12-month window (from {self._get_window_start_date()})")
    
    def _get_window_start_date(self) -> datetime:
        """Get the start date for rolling 12-month window (12 months ago from today)"""
        today = datetime.now()
        # Go back 11 months and start from the 1st of that month
        # This gives us 12 months total (current month + 11 previous months)
        # Calculate by subtracting months manually
        target_year = today.year
        target_month = today.month - 11
        
        # Handle year rollover
        while target_month <= 0:
            target_month += 12
            target_year -= 1
        
        start_date = datetime(target_year, target_month, 1)
        return start_date
    
    def _is_date_in_rolling_window(self, date: datetime) -> bool:
        """Check if a date falls within the rolling 12-month window"""
        if not self.use_rolling_window:
            return True  # No filtering if rolling window disabled
        
        window_start = self._get_window_start_date()
        today = datetime.now()
        
        # Include dates from window_start to today (inclusive)
        return window_start <= date <= today
    
    def _month_key_in_rolling_window(self, month_key: str) -> bool:
        """
        Check if a month key (format: "2025-一月") falls within rolling window.
        A month is included if ANY day in that month falls within the window.
        """
        if not self.use_rolling_window:
            return True
        
        try:
            # Parse month key like "2025-一月" or "2026-二月"
            year_str, month_cn = month_key.split('-', 1)
            year = int(year_str)
            
            # Map Chinese month to number
            month_map = {
                '一月': 1, '二月': 2, '三月': 3, '四月': 4, '五月': 5, '六月': 6,
                '七月': 7, '八月': 8, '九月': 9, '十月': 10, '十一月': 11, '十二月': 12
            }
            month = month_map.get(month_cn)
            
            if not month:
                return False
            
            # Check if the month overlaps with rolling window
            # Use first day of month and last day of month
            first_day = datetime(year, month, 1)
            # Get last day of month
            if month == 12:
                last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
            else:
                last_day = datetime(year, month + 1, 1) - timedelta(days=1)
            
            window_start = self._get_window_start_date()
            today = datetime.now()
            
            # Month is in window if it overlaps with [window_start, today]
            return first_day <= today and last_day >= window_start
            
        except (ValueError, KeyError):
            return False
    
    def get_rolling_12_months(self) -> Dict[str, pd.DataFrame]:
        """
        Get data for rolling 12-month window (last 12 months from today).
        This method filters the loaded data to only include months within the window.
        
        Returns:
            Dictionary of month keys to DataFrames, filtered to rolling 12 months
        """
        # Load all data first
        all_data = self.load_all_data(force_reload=False)
        
        # Filter to rolling 12-month window
        rolling_data = {}
        for month_key, df in all_data.items():
            if self._month_key_in_rolling_window(month_key):
                # Also filter transactions within each month by date
                if len(df) > 0 and 'date' in df.columns:
                    df_filtered = df[df['date'].apply(self._is_date_in_rolling_window)]
                    if len(df_filtered) > 0 or month_key.split('-')[1] in self._get_recent_months_list():
                        # Include month if it has data OR if it's one of the recent months in window
                        rolling_data[month_key] = df_filtered
                else:
                    # Empty dataframe, but month is in window
                    if self._month_key_in_rolling_window(month_key):
                        rolling_data[month_key] = df
        
        return rolling_data
    
    def _get_recent_months_list(self) -> List[str]:
        """Get list of Chinese month names in the rolling window"""
        today = datetime.now()
        window_start = self._get_window_start_date()
        
        months_list = []
        month_map = {
            1: '一月', 2: '二月', 3: '三月', 4: '四月', 5: '五月', 6: '六月',
            7: '七月', 8: '八月', 9: '九月', 10: '十月', 11: '十一月', 12: '十二月'
        }
        
        current = window_start
        while current <= today:
            months_list.append(month_map[current.month])
            # Move to next month
            if current.month == 12:
                current = datetime(current.year + 1, 1, 1)
            else:
                current = datetime(current.year, current.month + 1, 1)
        
        return months_list
    
    def load_all_data(self, force_reload: bool = False, use_rolling_window: Optional[bool] = None) -> Dict[str, pd.DataFrame]:
        """
        Load all months from all budget files and merge.
        If use_rolling_window is True, filters to rolling 12-month window.
        
        Args:
            force_reload: Force reload from files
            use_rolling_window: Override default rolling window setting (None = use self.use_rolling_window)
        
        Returns:
            Dictionary of month keys to DataFrames
        """
        # Determine if we should use rolling window
        should_use_rolling = use_rolling_window if use_rolling_window is not None else self.use_rolling_window
        
        # Always load raw data first
        all_data = self._load_all_data_raw(force_reload)
        
        # If rolling window is enabled, filter the data
        if should_use_rolling:
            filtered_data = {}
            for month_key, df in all_data.items():
                if self._month_key_in_rolling_window(month_key):
                    # Filter transactions by date within the month
                    if len(df) > 0 and 'date' in df.columns:
                        df_filtered = df[df['date'].apply(self._is_date_in_rolling_window)]
                        filtered_data[month_key] = df_filtered
                    else:
                        # Empty dataframe, but month is in window
                        filtered_data[month_key] = df
            
            return filtered_data
        
        # Return all data without filtering
        return all_data
    
    def _load_all_data_raw(self, force_reload: bool = False) -> Dict[str, pd.DataFrame]:
        """Internal method to load all data without filtering"""
        # Check cache
        if not force_reload and self._is_cache_valid():
            return self.cache
        
        print(f"📊 Loading multi-year budget data ({', '.join(map(str, self.years))})...")
        
        all_data = {}
        total_transactions = 0
        
        for budget_file, year in zip(self.budget_files, self.years):
            try:
                wb = load_workbook(budget_file, read_only=True, data_only=True)
                months = ['一月', '二月', '三月', '四月', '五月', '六月',
                         '七月', '八月', '九月', '十月', '十一月', '十二月']
                
                year_month_count = 0
                year_transaction_count = 0
                
                for month in months:
                    if month in wb.sheetnames:
                        ws = wb[month]
                        
                        rows = []
                        categories = ['交通費', '伙食費', '休閒/娛樂', '家務', '阿幫', '其它']
                        category_cols = [3, 4, 5, 6, 7, 8]
                        
                        for row in ws.iter_rows(min_row=3, values_only=True):
                            if row[0]:
                                date = row[0]
                                
                                # Convert string dates to datetime if needed
                                if isinstance(date, str):
                                    # Skip summary rows
                                    if any(keyword in date for keyword in 
                                        ['周總額', '單項總額', '月總額', '總計', '年度明細', '周总额', '单项总额']):
                                        continue
                                    # Try to parse date string
                                    try:
                                        from datetime import datetime as dt
                                        date = dt.strptime(date, '%Y-%m-%d')
                                    except (ValueError, TypeError):
                                        # Try other date formats or skip
                                        continue
                                elif not isinstance(date, datetime):
                                    # Not a datetime and not a parseable string, skip
                                    continue
                                else:
                                    # Already a datetime object, check for summary rows
                                    date_str = str(date)
                                    if any(keyword in date_str for keyword in 
                                        ['周總額', '單項總額', '月總額', '總計', '年度明細', '周总额', '单项总额']):
                                        continue
                                
                                # Extract each category amount
                                for cat, col_idx in zip(categories, category_cols):
                                    amount = row[col_idx] if col_idx < len(row) else None
                                    
                                    if amount and isinstance(amount, (int, float)) and amount > 0:
                                        rows.append({
                                            'date': date,
                                            'category': cat,
                                            'description': '',
                                            'amount': float(amount),
                                            'person': '',
                                            'year': year  # Track source year
                                        })
                        
                        # Always include the month if sheet exists, even if empty
                        # This ensures all months are visible even before data is added
                        if rows:
                            df = pd.DataFrame(rows)
                        else:
                            # Create empty DataFrame for months with no transactions yet
                            df = pd.DataFrame(columns=['date', 'category', 'description', 'amount', 'person', 'year'])
                        
                        # Key format: "2025-一月" or "2026-二月"
                        key = f"{year}-{month}"
                        all_data[key] = df
                        year_month_count += 1
                        year_transaction_count += len(rows)
                
                wb.close()
                
                if year_month_count > 0:
                    print(f"  ✅ {year}: Loaded {year_month_count} months with {year_transaction_count} transactions")
                    total_transactions += year_transaction_count
                
            except FileNotFoundError:
                print(f"  ⚠️  {year}: File not found (skipping)")
                continue
            except Exception as e:
                print(f"  ⚠️  {year}: Could not load ({e})")
                continue
        
        # Update cache
        self.cache = all_data
        self.last_loaded = datetime.now()
        
        print(f"✅ Total: {len(all_data)} months with {total_transactions} transactions")
        
        return all_data

    # ------------------------------------------------------------------
    # New helper methods for structured summaries
    # ------------------------------------------------------------------

    def _normalize_month_name(self, month: str) -> Optional[str]:
        """Normalize various month inputs to Chinese month names."""
        if not month:
            return None

        month = str(month).strip()

        # Already a composite key like "2025-七月"
        if '-' in month:
            _, suffix = month.split('-', 1)
            if suffix in self._month_names:
                return suffix

        lower = month.lower()

        # Exact Chinese month name (case-insensitive check)
        for cn in self._month_names:
            if cn in month:
                return cn

        # English month names
        if lower in self._english_month_map:
            return self._english_month_map[lower]

        # Numeric formats ("7", "07", "7月")
        digit = ''.join(ch for ch in month if ch.isdigit())
        if digit:
            try:
                idx = int(digit)
                if 1 <= idx <= 12:
                    return self._month_names[idx - 1]
            except ValueError:
                pass

        return None

    def _get_all_month_keys(self) -> List[str]:
        """Return all month keys (YYYY-月份) available across loaded years."""
        data = self._load_all_data_raw(force_reload=False)
        return list(data.keys())

    def resolve_month_key(self, month: str, year: Optional[int] = None) -> Optional[str]:
        """Resolve user-supplied month (with optional year) to stored month key."""
        if not month:
            return None

        all_keys = self._get_all_month_keys()
        if not all_keys:
            return None

        # Direct match
        if month in all_keys:
            return month

        month_name = self._normalize_month_name(month)
        if not month_name:
            return None

        candidates = []
        for key in all_keys:
            try:
                key_year_str, key_month = key.split('-', 1)
                key_year = int(key_year_str)
            except (ValueError, AttributeError):
                continue

            if key_month != month_name:
                continue

            if year is None or key_year == year:
                candidates.append((key_year, key))

        if not candidates:
            return None

        # Return the latest year match by default
        candidates.sort()
        return candidates[-1][1]

    def build_monthly_rollup(self, *,
                             use_rolling_window: bool = True,
                             force_reload: bool = False,
                             limit_months: Optional[int] = None) -> Dict[str, Dict[str, Dict]]:
        """
        Build structured summary for months within the rolling window.

        Returns:
            {
                'by_month': {
                    '2025-七月': {
                        'year': 2025,
                        'month': '七月',
                        'total': 85211.0,
                        'categories': {'伙食费': 30211.0, ...},
                        'top_categories': [('伙食费', 30211.0), ...],
                    },
                    ...
                },
                'month_order': ['2025-七月', '2025-八月', ...],
                'rolling_totals': {
                    'average_total': 91234.0,
                    'category_averages': {'伙食费': 28500.0, ...}
                }
            }
        """

        data = self.load_all_data(force_reload=force_reload, use_rolling_window=use_rolling_window)

        month_keys = sorted(data.keys())
        if limit_months and limit_months > 0:
            month_keys = month_keys[-limit_months:]

        by_month: Dict[str, Dict] = {}
        category_totals: Dict[str, float] = {}
        totals: List[float] = []
        months_with_data: List[str] = []

        for key in month_keys:
            df = data.get(key)

            if df is None or not hasattr(df, 'columns'):
                # Ensure we still register the sheet even if something unexpected happens
                df = pd.DataFrame(columns=['date', 'category', 'description', 'amount', 'person', 'year'])

            try:
                year_part, month_part = key.split('-', 1)
                year_value = int(year_part)
            except (ValueError, AttributeError):
                year_value = None
                month_part = key

            has_amount_col = 'amount' in df.columns
            has_category_col = 'category' in df.columns
            transaction_count = int(len(df)) if df is not None else 0
            has_data = transaction_count > 0 and has_amount_col

            total = float(df['amount'].sum()) if has_amount_col else 0.0
            categories_dict: Dict[str, float] = {}
            top_categories: List[Tuple[str, float]] = []

            if has_data and has_category_col:
                categories_series = df.groupby('category')['amount'].sum()
                categories_dict = {cat: float(amount) for cat, amount in categories_series.items()}
                top_categories = sorted(categories_dict.items(), key=lambda item: item[1], reverse=True)[:3]

                # Track category rollups for averages only when data exists
                for cat, amount in categories_dict.items():
                    category_totals[cat] = category_totals.get(cat, 0.0) + float(amount)

                totals.append(total)
                months_with_data.append(key)

            by_month[key] = {
                'year': year_value,
                'month': month_part,
                'total': total if has_amount_col else 0.0,
                'categories': categories_dict,
                'top_categories': top_categories,
                'transaction_count': transaction_count,
                'has_data': has_data,
                'has_amounts': has_amount_col,
            }

        average_total = sum(totals) / len(totals) if totals else 0.0

        category_averages = {
            cat: amount / len(months_with_data) for cat, amount in category_totals.items()
        } if months_with_data else {}

        return {
            'by_month': by_month,
            'month_order': month_keys,
            'months_with_data': months_with_data,
            'months_without_data': [key for key in month_keys if key not in months_with_data],
            'rolling_totals': {
                'average_total': average_total,
                'category_averages': category_averages,
                'months_counted': len(months_with_data)
            }
        }

    def compare_months(self,
                       month_a: str,
                       month_b: str,
                       *,
                       use_rolling_window: bool = True) -> Optional[Dict[str, Dict]]:
        """Return comparison metrics (totals + category deltas) between two months."""

        key_a = self.resolve_month_key(month_a)
        key_b = self.resolve_month_key(month_b)

        if not key_a or not key_b:
            return None

        rollup = self.build_monthly_rollup(use_rolling_window=use_rolling_window)
        month_data = rollup['by_month']

        data_a = month_data.get(key_a)
        data_b = month_data.get(key_b)

        if not data_a or not data_b:
            return None

        categories = set(data_a['categories'].keys()) | set(data_b['categories'].keys())
        category_deltas: Dict[str, Dict[str, float]] = {}

        for cat in categories:
            val_a = float(data_a['categories'].get(cat, 0.0))
            val_b = float(data_b['categories'].get(cat, 0.0))
            delta = val_b - val_a
            category_deltas[cat] = {
                'month_a': val_a,
                'month_b': val_b,
                'delta': delta
            }

        total_delta = float(data_b['total']) - float(data_a['total'])

        return {
            'month_a': {**data_a, 'key': key_a},
            'month_b': {**data_b, 'key': key_b},
            'total_delta': total_delta,
            'category_deltas': category_deltas
        }
    
    def get_summary_stats(self) -> Dict:
        """Get summary statistics for rolling 12-month window"""
        # Use rolling window by default
        data = self.load_all_data(force_reload=True, use_rolling_window=True)
        
        stats = {
            'total_months': len(data),
            'total_transactions': sum(len(df) for df in data.values()),
            'total_spending': 0,
            'by_category': {},
            'by_month': {},
            'by_year': {}  # Year breakdown
        }
        
        for month_key, df in data.items():
            if 'amount' in df.columns:
                month_total = df['amount'].sum()
                stats['total_spending'] += month_total
                stats['by_month'][month_key] = month_total
                
                # Year tracking
                year = df['year'].iloc[0] if 'year' in df.columns and len(df) > 0 else None
                if year:
                    stats['by_year'][year] = stats['by_year'].get(year, 0) + month_total
                
                # Category breakdown
                if 'category' in df.columns:
                    for cat, amount in df.groupby('category')['amount'].sum().items():
                        stats['by_category'][cat] = stats['by_category'].get(cat, 0) + amount
        
        return stats
    
    def get_year_range(self) -> tuple:
        """
        Return (min_year, max_year)
        
        Returns:
            Tuple of (earliest_year, latest_year) or (None, None) if no data
        """
        if self.years:
            return (min(self.years), max(self.years))
        return (None, None)
    
    def get_available_months(self) -> List[str]:
        """
        Get list of available months in format 'YYYY-月份' (rolling 12-month window)
        
        Returns:
            Sorted list of month keys within rolling window
        """
        # Use rolling window by default
        data = self.load_all_data(force_reload=True, use_rolling_window=True)
        return sorted(data.keys())

