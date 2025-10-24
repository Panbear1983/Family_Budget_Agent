"""
Data Loader - Efficiently loads and caches budget data
"""

import pandas as pd
from openpyxl import load_workbook
from typing import Dict, List, Optional
from datetime import datetime

class DataLoader:
    """Loads and caches budget data from Excel"""
    
    def __init__(self, budget_file: str):
        self.budget_file = budget_file
        self.cache = {}
        self.last_loaded = None
        self.ttl = 1800  # Cache for 30 minutes
    
    def load_all_data(self, force_reload: bool = False, silent: bool = False) -> Dict[str, pd.DataFrame]:
        """Load all months from budget file"""
        
        # Check cache
        if not force_reload and self._is_cache_valid():
            return self.cache
        
        if not silent:
            print("📊 Loading budget data...")
        
        try:
            wb = load_workbook(self.budget_file, read_only=True, data_only=True)
            months = ['一月', '二月', '三月', '四月', '五月', '六月',
                     '七月', '八月', '九月', '十月', '十一月', '十二月']
            
            data = {}
            for month in months:
                if month in wb.sheetnames:
                    ws = wb[month]
                    
                    # Convert wide format to long format
                    rows = []
                    
                    # Categories are in columns D-I (index 3-8)
                    categories = ['交通費', '伙食費', '休閒/娛樂', '家務', '阿幫', '其它']
                    category_cols = [3, 4, 5, 6, 7, 8]  # Column indices
                    
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        if row[0]:  # If date exists
                            date = row[0]
                            
                            # SKIP SUMMARY ROWS - Only process actual datetime objects
                            if not isinstance(date, datetime):
                                continue
                            
                            # FILTER BY YEAR - Only include 2025 data for 2025 reports
                            # Note: Year filtering removed for performance - only 2025 data exists
                            # if date.year != 2025:
                            #     continue
                            
                            # SKIP rows with summary keywords (extra safety)
                            date_str = str(date)
                            if any(keyword in date_str for keyword in 
                                ['周總額', '單項總額', '月總額', '總計', '年度明細', '周总额', '单项总额']):
                                continue
                            
                            # Extract each category amount
                            for cat, col_idx in zip(categories, category_cols):
                                amount = row[col_idx] if col_idx < len(row) else None
                                
                                if amount and isinstance(amount, (int, float)) and amount > 0:
                                    rows.append({
                                        'date': date,
                                        'category': cat,
                                        'description': '',
                                        'amount': float(amount),
                                        'person': ''
                                    })
                    
                    if rows:
                        df = pd.DataFrame(rows)
                        data[month] = df
            
            wb.close()
            
            # Update cache
            self.cache = data
            self.last_loaded = datetime.now()
            
            if not silent:
                print(f"✅ Loaded {len(data)} months with {sum(len(df) for df in data.values())} transactions")
            return data
        
        except Exception as e:
            print(f"❌ Error loading data: {e}")
            import traceback
            traceback.print_exc()
            return {}
    
    def load_month(self, month: str) -> Optional[pd.DataFrame]:
        """Load specific month"""
        all_data = self.load_all_data()
        return all_data.get(month)
    
    def get_summary_stats(self, silent: bool = False) -> Dict:
        """Get quick summary statistics"""
        data = self.load_all_data(silent=silent)
        
        stats = {
            'total_months': len(data),
            'total_transactions': sum(len(df) for df in data.values()),
            'total_spending': 0,
            'by_category': {},
            'by_month': {}
        }
        
        for month, df in data.items():
            if 'amount' in df.columns:
                month_total = df['amount'].sum()
                stats['total_spending'] += month_total
                stats['by_month'][month] = month_total
                
                # Category breakdown
                if 'category' in df.columns:
                    for cat, amount in df.groupby('category')['amount'].sum().items():
                        stats['by_category'][cat] = stats['by_category'].get(cat, 0) + amount
        
        return stats
    
    def clear_cache(self, silent: bool = False):
        """Clear the data cache to force fresh data loading"""
        self.cache = {}
        self.last_loaded = None
        if not silent:
            print("🔄 Cache cleared - will reload data with year filtering")
    
    def _is_cache_valid(self) -> bool:
        """Check if cache is still valid"""
        if not self.cache or not self.last_loaded:
            return False
        
        elapsed = (datetime.now() - self.last_loaded).total_seconds()
        return elapsed < self.ttl

