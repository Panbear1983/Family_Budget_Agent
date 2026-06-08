"""
Monthly Merger - Merge your + wife's exports into main budget
"""

import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from core.base_module import BaseModule
from utils.excel_totals import recalculate_month_totals
from typing import Tuple

class MonthlyMerger(BaseModule):
    """Handle monthly budget merges"""
    
    def _setup(self):
        """Initialize merger"""
        self.categorizer = None  # Will be set by orchestrator
        self.orchestrator = None
        print("  📥 Monthly Merger initialized")
    
    def set_categorizer(self, categorizer):
        """Set categorizer module"""
        self.categorizer = categorizer
    
    def set_orchestrator(self, orchestrator):
        """Set orchestrator for LLM access"""
        self.orchestrator = orchestrator
    
    def execute_from_dataframes(self, peter_df: pd.DataFrame, dolly_df: pd.DataFrame, 
                                target_month: str, main_budget_path: str, merge_mode: bool = False):
        """
        Execute merge from pre-parsed DataFrames (from FileParser)
        Returns: (success, merged_count, merged_dataframe)
        """
        try:
            print(f"\n🔍 Categorizing transactions...")
            
            # Categorize Peter's transactions
            if len(peter_df) > 0:
                peter_categorized = self.categorizer.batch_categorize(
                    peter_df.to_dict('records'),
                    person='peter'
                )
            else:
                peter_categorized = []
            
            # Categorize Dolly's transactions
            if len(dolly_df) > 0:
                dolly_categorized = self.categorizer.batch_categorize(
                    dolly_df.to_dict('records'),
                    person='wife'
                )
            else:
                dolly_categorized = []
            
            # Combine
            all_transactions = peter_categorized + dolly_categorized
            combined_df = pd.DataFrame(all_transactions)
            print(f"  ✅ Combined: {len(combined_df)} total transactions")
            
            # Deduplicate
            print(f"\n🔍 Detecting duplicates...")
            deduped = self.deduplicate(combined_df)
            print(f"  ✅ After dedup: {len(deduped)} unique transactions")
            
            # Preview
            self.show_preview(deduped, target_month)
            
            return True, len(deduped), deduped
            
        except Exception as e:
            return False, 0, f"Error: {str(e)}"
    
    def execute(self, *args, **kwargs):
        """Deprecated — use execute_from_dataframes() with pre-parsed FileParser DataFrames."""
        raise NotImplementedError(
            "execute() is deprecated and reads from row 1, producing garbage output. "
            "Use execute_from_dataframes() with DataFrames from FileParser instead."
        )
    
    def deduplicate(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Pass-through — no deduplication performed.
        MonnyReport export is the source of truth; all rows are written as-is.
        Data accuracy is the user's responsibility at the source (MonnyReport app).
        """
        df['date'] = pd.to_datetime(df['date'], errors='coerce')
        return df
    
    def show_preview(self, df: pd.DataFrame, month: str):
        """Show detailed preview in annual sheet format"""
        print("\n" + "="*100)
        print(f"  📋 PREVIEW: {month} Sheet (As it will appear)".center(100))
        print("="*100 + "\n")
        
        # Summary stats
        print("📊 SUMMARY:")
        print(f"   Total Transactions: {len(df)}")
        print(f"   Total Amount: NT${df['amount'].sum():,.0f}")
        print(f"   Date Range: {df['date'].min().date()} to {df['date'].max().date()}")
        
        # By category
        print("\n📁 BY CATEGORY:")
        by_cat = df.groupby('main_category')['amount'].agg(['count', 'sum'])
        for cat, row in by_cat.iterrows():
            print(f"   {cat:15s}: {int(row['count']):3d} items = NT${int(row['sum']):,}")
        
        # Group by date for daily entries
        daily = df.groupby(df['date'].dt.date).apply(
            lambda x: x.groupby('main_category')['amount'].sum().to_dict()
        ).to_dict()
        
        # Table header (matches annual sheet format)
        print("\n" + "="*100)
        print(f"{'日期':12s} │ {'星期':6s} │ {'交通費':>8s} │ {'伙食費':>8s} │ {'休閒/娛樂':>10s} │ {'家務':>6s} │ {'阿幫':>6s} │ {'其它':>6s} │ {'每日總額':>10s}")
        print("─" * 100)
        
        # Weekly totals tracking
        weekly_totals = {
            '交通費': 0, '伙食費': 0, '休閒/娛樂': 0,
            '家務': 0, '阿幫': 0, '其它': 0
        }
        weekly_total = 0
        last_week = None
        
        # Data rows
        for date, cats in sorted(daily.items()):
            dow = pd.Timestamp(date).strftime('%A')
            dow_chinese = {
                'Monday': '星期一', 'Tuesday': '星期二', 'Wednesday': '星期三',
                'Thursday': '星期四', 'Friday': '星期五', 
                'Saturday': '星期六', 'Sunday': '星期天'
            }.get(dow, dow)
            
            # Check if we need to print weekly total
            current_week = pd.Timestamp(date).isocalendar()[1]
            if last_week is not None and current_week != last_week and weekly_total > 0:
                # Print weekly total row
                print("─" * 100)
                wt_trans = weekly_totals['交通費']
                wt_food = weekly_totals['伙食費']
                wt_fun = weekly_totals['休閒/娛樂']
                wt_home = weekly_totals['家務']
                wt_helper = weekly_totals['阿幫']
                wt_other = weekly_totals['其它']
                print(f"{'周總額':12s} │ {'      ':6s} │ {wt_trans:8.0f} │ {wt_food:8.0f} │ {wt_fun:10.0f} │ {wt_home:6.0f} │ {wt_helper:6.0f} │ {wt_other:6.0f} │ {weekly_total:10.0f}")
                print("─" * 100)
                
                # Reset weekly totals
                weekly_totals = {k: 0 for k in weekly_totals}
                weekly_total = 0
            
            last_week = current_week
            
            # Print daily row
            trans = cats.get('交通費', 0)
            food = cats.get('伙食費', 0)
            fun = cats.get('休閒/娛樂', 0)
            home = cats.get('家務', 0)
            helper = cats.get('阿幫', 0)
            other = cats.get('其它', 0)
            total = trans + food + fun + home + helper + other
            
            print(f"{date} │ {dow_chinese:6s} │ {trans:8.0f} │ {food:8.0f} │ {fun:10.0f} │ {home:6.0f} │ {helper:6.0f} │ {other:6.0f} │ {total:10.0f}")
            
            # Accumulate weekly totals
            weekly_totals['交通費'] += trans
            weekly_totals['伙食費'] += food
            weekly_totals['休閒/娛樂'] += fun
            weekly_totals['家務'] += home
            weekly_totals['阿幫'] += helper
            weekly_totals['其它'] += other
            weekly_total += total
        
        # Print final weekly total if any data remains
        if weekly_total > 0:
            print("─" * 100)
            wt_trans = weekly_totals['交通費']
            wt_food = weekly_totals['伙食費']
            wt_fun = weekly_totals['休閒/娛樂']
            wt_home = weekly_totals['家務']
            wt_helper = weekly_totals['阿幫']
            wt_other = weekly_totals['其它']
            print(f"{'周總額':12s} │ {'      ':6s} │ {wt_trans:8.0f} │ {wt_food:8.0f} │ {wt_fun:10.0f} │ {wt_home:6.0f} │ {wt_helper:6.0f} │ {wt_other:6.0f} │ {weekly_total:10.0f}")
        
        print("\n" + "="*100)
    
    def wipe_month_tab(self, month_name: str, budget_file: str) -> bool:
        """
        Fully wipe the month input grid (safe wipe).

        Clears only the user-input amount cells in the daily calendar blocks:
        - Rows: 3-9, 11-17, 19-25, 27-33, 35-41, 43-49
        - Cols: D-I (4-9) (交通費..其它)

        Does NOT touch date/weekday (A/B) or any formulas/labels outside this grid.
        """
        try:
            from openpyxl import load_workbook

            wb = load_workbook(budget_file)
            if month_name not in wb.sheetnames:
                print(f"  ❌ Sheet '{month_name}' not found in budget file")
                return False

            ws = wb[month_name]

            weekly_blocks = [
                (3, 9),
                (11, 17),
                (19, 25),
                (27, 33),
                (35, 41),
                (43, 49),
            ]

            cleared = 0
            for start_row, end_row in weekly_blocks:
                for row_num in range(start_row, end_row + 1):
                    for col_num in range(4, 10):  # D(4) .. I(9)
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            cell.value = None
                            cleared += 1

            wb.save(budget_file)
            print(f"  🧹 Wiped {month_name} input grid: cleared {cleared} cell(s)")
            return True
        except Exception as e:
            print(f"  ❌ Wipe failed: {e}")
            return False

    def append_to_month_tab(self, df: pd.DataFrame, month_name: str, budget_file: str, merge_mode: bool = False):
        """
        Write transactions to monthly tab with calendar-aligned row placement.
        Only writes to restricted areas: rows 3-9, 11-17, 19-25, 27-33, 35-41, 43-49
        Only writes amounts to columns D-I (4-9).
        Assumes columns A & B (date & weekday) are pre-filled by auto-fill function.

        Args:
            df: DataFrame with transactions
            month_name: Name of the month sheet
            budget_file: Path to budget Excel file
            merge_mode: If True, add to existing values; if False, overwrite (clears stale values)
        """
        # Load read-only copy with cached cell values for date scanning (handles formula cells)
        wb_ro = load_workbook(budget_file, data_only=True)
        if month_name not in wb_ro.sheetnames:
            print(f"  ❌ Sheet '{month_name}' not found in budget file")
            return False
        ws_ro = wb_ro[month_name]

        # Load writable copy for writing amounts
        wb = load_workbook(budget_file)
        ws = wb[month_name]

        # Define weekly blocks (Mon-Sun rows for each week)
        weekly_blocks = [
            (3, 9),    # Week 1: Mon(3), Tue(4), Wed(5), Thu(6), Fri(7), Sat(8), Sun(9)
            (11, 17),  # Week 2: Mon(11) through Sun(17)
            (19, 25),  # Week 3
            (27, 33),  # Week 4
            (35, 41),  # Week 5
            (43, 49),  # Week 6
        ]

        # Column mapping (only write to columns D-I for amounts)
        col_map = {
            '交通費': 4,      # Column D
            '伙食費': 5,      # Column E
            '休閒/娛樂': 6,   # Column F
            '家務': 7,        # Column G
            '阿幫': 8,        # Column H
            '其它': 9         # Column I
        }

        # Group by date for daily entries
        daily_totals = df.groupby(df['date'].dt.date).apply(
            lambda x: x.groupby('main_category')['amount'].sum().to_dict()
        ).to_dict()

        mode_text = "merge mode (adding to existing)" if merge_mode else "overwrite mode (replacing existing)"
        print(f"\n  📅 Writing {len(daily_totals)} days to {month_name} (calendar-aligned, {mode_text})")

        # Build date-to-row mapping by reading cached date values from column A (read-only copy)
        print("  🔍 Reading dates from Excel to find correct row positions...")
        date_to_row = {}

        for row_num in range(3, 50):
            date_cell = ws_ro.cell(row_num, 1).value  # Column A — cached value, not formula string
            if not date_cell:
                continue
            date_key = None
            if isinstance(date_cell, datetime):
                date_key = date_cell.date()
            else:
                try:
                    date_key = pd.to_datetime(date_cell).date()
                except Exception:
                    continue
            if date_key:
                date_to_row[date_key] = row_num

        print(f"  ✅ Found {len(date_to_row)} dates in Excel sheet")

        # Fallback: if no dates found (sheet not yet filled), derive from calendar
        if len(date_to_row) == 0:
            import calendar
            print("  ⚠️  No dates found in column A; using calendar fallback mapping")
            calendar_cache = {}
            for date_key in sorted(daily_totals.keys()):
                year, month = date_key.year, date_key.month
                cache_key = (year, month)
                if cache_key not in calendar_cache:
                    calendar_cache[cache_key] = calendar.monthrange(year, month)
                first_weekday, days_in_month = calendar_cache[cache_key]
                if date_key.day > days_in_month:
                    continue
                offset = first_weekday + (date_key.day - 1)
                week_index = offset // 7
                weekday_index = offset % 7
                if week_index >= len(weekly_blocks):
                    continue
                date_to_row[date_key] = weekly_blocks[week_index][0] + weekday_index
            print(f"  ✅ Generated {len(date_to_row)} dates from calendar fallback")

        # Process each date
        written_count = 0
        skipped_count = 0

        for date, categories in sorted(daily_totals.items()):
            row_num = date_to_row.get(date)

            if row_num is None:
                print(f"  ⚠️  Date {date} not found in Excel sheet (skipping)")
                skipped_count += 1
                continue

            # Safety check: row must be inside a valid weekly block
            in_valid_block = any(s <= row_num <= e for s, e in weekly_blocks)
            if not in_valid_block:
                print(f"  ⚠️  Date {date} at row {row_num} outside valid weekly blocks (skipping)")
                skipped_count += 1
                continue

            # Write amounts to columns D-I (4-9)
            for cat, col in col_map.items():
                amount = categories.get(cat, 0)
                if merge_mode:
                    if amount > 0:
                        existing = ws.cell(row_num, col).value or 0
                        ws.cell(row_num, col, existing + amount)
                else:
                    # Overwrite mode: always write — zero clears any stale value from a prior merge
                    ws.cell(row_num, col, amount if amount > 0 else None)

            written_count += 1

        recalculate_month_totals(ws)

        wb.calculation.calcMode = 'auto'
        wb.save(budget_file)

        if skipped_count > 0:
            print(f"  ⚠️  Skipped {skipped_count} date(s) outside 6-week range")

        print(f"  ✅ Updated {month_name} tab with {written_count} daily entries")
        print(f"  ☁️  OneDrive will auto-sync changes")

        return True

