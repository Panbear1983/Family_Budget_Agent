"""
Annual Manager - Handle annual budget file creation
"""

import os
import shutil
from datetime import datetime
from openpyxl import load_workbook, Workbook
from core.base_module import BaseModule

class AnnualManager(BaseModule):
    """Manage annual budget file lifecycle"""
    
    def _setup(self):
        """Initialize annual manager"""
        self.onedrive_path = self.config.get('onedrive_path', '')
        self.template_file = self.config.get('template_file', 'TEMPLATE_年開銷表.xlsx')
        self.auto_create = self.config.get('auto_create', True)
        print("  📅 Annual Manager initialized")
    
    def execute(self, year: int = None):
        """
        Check and create annual budget file if needed
        Returns: (file_path, created)
        """
        if year is None:
            year = datetime.now().year
        
        budget_file = self.get_budget_file_path(year)
        
        if os.path.exists(budget_file):
            return budget_file, False  # Already exists
        
        if self.auto_create:
            print(f"\n🆕 Creating budget file for {year}...")
            self.create_annual_budget(year)
            return budget_file, True
        
        return None, False
    
    def get_budget_file_path(self, year: int) -> str:
        """Get path for year's budget file"""
        filename = f"{year}年開銷表（NT）.xlsx"
        
        if self.onedrive_path:
            return os.path.join(self.onedrive_path, filename)
        else:
            return filename
    
    def get_active_budget_file(self) -> str:
        """Get current year's budget file, create if needed"""
        current_year = datetime.now().year
        budget_file, created = self.execute(current_year)
        
        if created:
            print(f"  ✅ Created new budget file for {current_year}")
        
        return budget_file
    
    def create_annual_budget(self, year: int):
        """
        Create new annual budget file with auto-filled dates
        Priority: Template > Clone previous > Create new
        """
        target_file = self.get_budget_file_path(year)
        
        # Option 1: Use template if exists
        if os.path.exists(self.template_file):
            print(f"  📋 Using template: {self.template_file}")
            shutil.copy2(self.template_file, target_file)
            print(f"  ✅ Created from template: {target_file}")
            # Auto-fill dates for all months
            self.auto_fill_all_dates(target_file, year)
            return target_file
        
        # Option 2: Clone previous year
        prev_year_file = self.get_budget_file_path(year - 1)
        if os.path.exists(prev_year_file):
            print(f"  📋 Cloning structure from {year - 1}")
            self.clone_and_clear(prev_year_file, target_file)
            print(f"  ✅ Created from {year - 1}: {target_file}")
            # Auto-fill dates for all months
            self.auto_fill_all_dates(target_file, year)
            return target_file
        
        # Option 3: Create from scratch
        print(f"  📋 Creating new structure from scratch")
        self.create_from_scratch(target_file)
        print(f"  ✅ Created new file: {target_file}")
        # Auto-fill dates for all months
        self.auto_fill_all_dates(target_file, year)
        return target_file
    
    def clone_and_clear(self, source_file: str, target_file: str):
        """
        Clone structure from previous year but clear all data
        Keeps formulas, formatting, structure
        """
        wb = load_workbook(source_file)
        
        # For each sheet (month)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Clear data cells (rows 3 onwards, but keep structure)
            # Keep row 1-2 (headers)
            # Clear rows 3-48 (daily entries area)
            for row in range(3, 49):
                for col in range(1, 12):
                    cell = ws.cell(row, col)
                    
                    # Keep if it's a label/structure cell
                    cell_val = str(cell.value or '')
                    if any(keyword in cell_val for keyword in ['週總額', '周總額', '單項總額', '年度明細', '星期']):
                        continue  # Keep structure
                    
                    # Clear data
                    if cell.value is not None:
                        # Keep formula if exists, otherwise clear
                        if not str(cell.value).startswith('='):
                            cell.value = None
            
            # Clear monthly summary rows (49-62)
            for row in range(49, 63):
                for col in range(3, 10):  # Clear amount columns only
                    cell = ws.cell(row, col)
                    if cell.value is not None and not str(cell.value).startswith('='):
                        cell.value = None
        
        wb.save(target_file)
    
    def create_from_scratch(self, target_file: str):
        """
        Create basic annual budget structure from scratch
        """
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Create 12 month sheets
        months = ['一月', '二月', '三月', '四月', '五月', '六月', 
                 '七月', '八月', '九月', '十月', '十一月', '十二月']
        
        for month in months:
            ws = wb.create_sheet(month)
            
            # Create headers
            ws.cell(1, 1, '日期：')
            ws.cell(1, 2, '星期:')
            ws.cell(1, 4, '交通费：')
            ws.cell(1, 5, '伙食费：')
            ws.cell(1, 6, '休闲/娱乐：')
            ws.cell(1, 7, '家务：')
            ws.cell(1, 8, '阿幫：')
            ws.cell(1, 9, '其它：')
            ws.cell(1, 11, '每日總額')
            
            # Add day of week labels
            days = ['星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期天']
            for i, day in enumerate(days, start=2):
                ws.cell(i, 2, day)
        
        wb.save(target_file)
    
    def auto_fill_all_dates(self, target_file: str, year: int):
        """
        Automatically fill dates for all 12 months in the new year file
        Uses datetime to determine what day of the week each month starts
        Follows exact same logic as edit_cells.py autofill_dates_workflow
        """
        import calendar
        
        months = ['一月', '二月', '三月', '四月', '五月', '六月', 
                 '七月', '八月', '九月', '十月', '十一月', '十二月']
        
        try:
            wb = load_workbook(target_file)
            
            for month_idx, month_name in enumerate(months, start=1):
                if month_name not in wb.sheetnames:
                    continue
                    
                ws = wb[month_name]
                
                # Auto-detect what day of the week this month starts
                first_date = datetime(year, month_idx, 1)
                start_weekday = first_date.isoweekday()  # 1=Mon, 7=Sun
                
                # Get number of days in this month
                days_in_month = calendar.monthrange(year, month_idx)[1]
                
                # Week blocks (EXACT SAME as edit_cells.py)
                week_blocks = [
                    (3, 9),    # Week 1: rows 3-9
                    (11, 17),  # Week 2
                    (19, 25),  # Week 3
                    (27, 33),  # Week 4
                    (35, 41),  # Week 5
                    (43, 49),  # Week 6
                ]
                
                # Fill dates (EXACT SAME logic as edit_cells.py)
                week_idx = 0
                day_idx = start_weekday - 1  # 0-6 (Mon=0, Sun=6)
                
                for day_num in range(1, days_in_month + 1):
                    date_str = f"{year}-{month_idx:02d}-{day_num:02d}"
                    
                    if week_idx < len(week_blocks):
                        start_row, end_row = week_blocks[week_idx]
                        row_idx = start_row - 1 + day_idx
                        
                        # Write date to column A (SAME as edit_cells.py)
                        ws.cell(row=row_idx + 1, column=1, value=date_str)
                    
                    # Move to next day (SAME logic)
                    day_idx += 1
                    if day_idx >= 7:  # New week
                        day_idx = 0
                        week_idx += 1
                
                print(f"  ✅ Auto-filled {days_in_month} dates for {month_name}")
            
            wb.save(target_file)
            print(f"  📅 All dates auto-filled for {year}!")
            
        except Exception as e:
            print(f"  ⚠️  Date auto-fill failed: {e}")
            print(f"     You can still manually fill dates using '填充日期' menu")
    
    def archive_old_year(self, year: int):
        """Move old year's budget to archive"""
        archive_dir = self.config.get('archive_path', 'archive')
        
        if not os.path.exists(archive_dir):
            os.makedirs(archive_dir)
        
        old_file = self.get_budget_file_path(year)
        if os.path.exists(old_file):
            archive_file = os.path.join(archive_dir, os.path.basename(old_file))
            shutil.move(old_file, archive_file)
            print(f"  📦 Archived {year} budget to {archive_file}")
    
    def get_multi_year_files(self, num_years: int = 2) -> list:
        """
        Get budget files for multiple years (current + previous)
        
        Args:
            num_years: Number of years to load (default 2: current + previous)
        
        Returns:
            List of file paths, sorted by year (oldest first)
        """
        current_year = datetime.now().year
        # Load current year and previous year, but exclude 2024 specifically
        years = list(range(current_year - num_years + 1, current_year + 1))
        # Filter out 2024 data specifically
        years = [year for year in years if year != 2024]
        
        files = []
        for year in years:
            file_path = self.get_budget_file_path(year)
            if os.path.exists(file_path):
                files.append(file_path)
            else:
                print(f"  ⚠️  {year} budget file not found (skipping)")
        
        return files

